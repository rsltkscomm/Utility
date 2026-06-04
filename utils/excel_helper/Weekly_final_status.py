from openpyxl import load_workbook
from datetime import datetime, timedelta
from pathlib import Path

from utils.constants import framework_constants
from utils.constants.framework_constants import FrameworkConstants
from utils.ini_file_reader.config_reader import ConfigReader


class WeeklyStatusManager:
    FILE_PATH = FrameworkConstants.ONEDRIVE_BASE_PATH / "Daily.xlsx"
    print(FILE_PATH)
    EXISTS = True
    print("FILE_PATH =", FILE_PATH)
    print("EXISTS =", FILE_PATH.exists())
    print("SUFFIX =", FILE_PATH.suffix)

    @staticmethod
    def read_last_7_days_status(FILE_PATH):
        try:
            suite_name = ConfigReader.get_property("SuiteName", "").strip()
        except Exception:
            suite_name = ""

        suite_name_lower = suite_name.lower()
        is_daily_suite = "daily" in suite_name_lower

        try:
            wb = load_workbook(FILE_PATH, read_only=True, data_only=True)
        except Exception as e:
            print(f"[WARNING] Could not open Excel file '{FILE_PATH}': {e}. History data will be empty.")
            return {}

        try:
            ws = wb.active

            headers = {}

            # Read headers from row 1.
            # Date columns may be stored as datetime objects in Excel (not strings).
            # Normalise date headers to YYYY-MM-DD; leave text headers (e.g.
            # "Module Name", "Test Case ID") untouched so lookups still work.
            import re as _re
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=1, column=col).value
                if value is None:
                    continue
                if hasattr(value, 'strftime'):
                    key = value.strftime("%Y-%m-%d")
                else:
                    raw = str(value).strip()
                    if _re.match(r'\d{4}-\d{2}-\d{2}', raw):
                        key = raw[:10]
                    else:
                        key = raw
                headers[key] = col

            module_col = headers.get("Module Name")
            testcase_col = headers.get("Test Case ID")

            if not module_col or not testcase_col:
                print(
                    f"[WARNING] Required columns 'Module Name' or 'Test Case ID' not found in history Excel. "
                    f"Found headers: {list(headers.keys())}. History data will be empty."
                )
                return {}

            # Generate last 7 dates oldest -> newest
            last_7_days = []
            for i in range(6, -1, -1):
                date = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
                last_7_days.append(date)

            result = {}

            for row in range(2, ws.max_row + 1):
                module_value = ws.cell(row=row, column=module_col).value

                if module_value is None:
                    continue

                module_value_str = str(module_value).strip()
                module_value_lower = module_value_str.lower()

                # ── Filtering logic based on SuiteName ──────────────────────────
                # If running a daily suite → match ONLY rows whose Module Name is
                # exactly the static marker 'dailychecklist--UPDATE_ZEPHYR_EXECUTION=yes'.
                # For any other suite → exact (case-insensitive) match against
                # SuiteName so only the rows belonging to that suite are included.
                if is_daily_suite:
                    if module_value_str != "dailychecklist--UPDATE_ZEPHYR_EXECUTION=yes":
                        continue
                else:
                    if module_value_lower != suite_name_lower:
                        continue
                # ────────────────────────────────────────────────────────────────

                testcase_value = ws.cell(row=row, column=testcase_col).value
                if not testcase_value:
                    continue

                testcase_value = str(testcase_value).strip()

                history = []
                for date in last_7_days:
                    date_col = headers.get(date)

                    if not date_col:
                        history.append("SKIPPED")
                        continue

                    status_value = ws.cell(row=row, column=date_col).value

                    if not status_value:
                        history.append("SKIPPED")
                    else:
                        history.append(str(status_value).strip().upper())

                result[testcase_value] = history

            print(f"[INFO] Loaded history for {len(result)} test cases from Excel "
                  f"(SuiteName='{suite_name}', daily_mode={is_daily_suite}).")
            return result

        except Exception as e:
            import traceback
            print(f"[WARNING] Error reading history data from '{FILE_PATH}': {e}. History data will be empty.")
            traceback.print_exc()
            return {}
        finally:
            try:
                wb.close()
            except Exception:
                pass